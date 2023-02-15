#!/usr/bin/python3.8
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import datetime
import telebot
import logging
from telebot import types
import array

joinedFile = open('data/ids.txt', 'r')
joinedUsers = set()
for line in joinedFile:
    joinedUsers.add(line.strip())
joinedFile.close()

wb = load_workbook('data/03.09.xlsx')
ws = wb.active

TOKEN = os.getenv("TOKEN")

time_str = datetime.datetime.today()
timeadd = time_str + datetime.timedelta(days=1)
current_time = time_str.strftime('%a')
day_plus_one = timeadd.strftime('%a')

logger = telebot.logger
telebot.logger.setLevel(logging.DEBUG)  # Outputs debug messages to console.

markup = types.ReplyKeyboardMarkup(True)
itembtn1 = types.KeyboardButton('/today')
itembtn2 = types.KeyboardButton('/week')
itembtn3 = types.KeyboardButton('/nextday')
itembtn4 = types.KeyboardButton('/Расписание_звонков')
markup.add(itembtn1, itembtn3, itembtn2, itembtn4)

bot = telebot.TeleBot(TOKEN)


@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    bot.reply_to(message, "Бот с расписанием, чтобы начать напишите  /start ", reply_markup=markup)
    if not str(message.chat.id) in joinedUsers:
        joinedFile = open('ids.txt', 'a')
        joinedFile.write(str(message.chat.id) + "\n")
        joinedUsers.add(message.chat.id)


monday_array = []

def mondays(message):
    global mondayt
    p = -1
    for row in range(4, 10):
        char = get_column_letter(2)
        monday = ws[char + str(row)].value
        if monday != None:
            monday_array.append(monday)
    if monday_array != None:
        print(monday_array)
        p += 1
        mondayt = f'{monday_array[p]}.\n{monday_array[p + 1]}.\n{monday_array[p + 2]}'
    bot.reply_to(message, mondayt)


tuesday_array = []

def tuesdays(message):
    global tuesdayt
    p = -1
    for row in range(4, 10):
        char = get_column_letter(4)
        tuesday = ws[char + str(row)].value
        if tuesday != None:
            tuesday_array.append(tuesday)
    if tuesday_array != None:
        print(tuesday_array)
        p += 1
        tuesdayt = f'{tuesday_array[p]}.\n{tuesday_array[p + 1]}.\n{tuesday_array[p + 2]}.\n{tuesday_array[p + 3]}'
    bot.reply_to(message, tuesdayt)


wednesday_array = []

def wednesdays(message):
    global wednesdayt
    p = -1
    for row in range(4, 10):
        char = get_column_letter(6)
        wednesday = ws[char + str(row)].value
        if wednesday != None:
            wednesday_array.append(wednesday)
    if wednesday_array != None:
        print(wednesday_array)
        p += 1
        wednesdayt = f'{wednesday_array[p]}.\n{wednesday_array[p + 1]}.\n{wednesday_array[p + 2]}'
    bot.send_message(message, wednesdayt)


thursday_array = []

def thursdays(message):
    global thursdayt
    p = -1
    for row in range(4, 10):
        char = get_column_letter(8)
        thursday = ws[char + str(row)].value
        if thursday != None:
            thursday_array.append(thursday)
    if thursday_array != None:
        print(thursday_array)
        p += 1
        thursdayt = f'{thursday_array[p]}.\n{thursday_array[p + 1]}.\n{thursday_array[p + 2]}'
    bot.reply_to(message, thursdayt)


friday_array = []

def fridays(message):
    global fridayt
    p = -1
    for row in range(4, 10):
        char = get_column_letter(10)
        friday = ws[char + str(row)].value
        if friday != None:
            friday_array.append(friday)
    if friday_array != None:
        print(friday_array)
        p += 1
        fridayt = f'{friday_array[p]}.\n{friday_array[p + 1]}.\n{friday_array[p + 2]}.\n{friday_array[p + 3]}'
    bot.reply_to(message, fridayt)


saturday_array = []

def saturdays(message):
    global saturdayt
    p = -1
    for row in range(4, 10):
        col = 12
        char = get_column_letter(col)
        saturday = ws[char + str(row)].value
        if saturday != None:
            saturday_array.append(saturday)
    if saturday_array != None:
        print(saturday_array)
        p += 1
        saturdayt = f'{saturday_array[p]}.\n{saturday_array[p]}'
    bot.reply_to(message, saturdayt)


@bot.message_handler(commands=['today'])
def send_text(message):

    err_gf = open('data/err.gif', 'rb')

    if current_time == 'Mon':
        mondays(message)
    elif current_time == 'Tue':
        tuesdays(message)
    elif current_time == 'Wed':
        wednesdays(message)
    elif current_time == 'Thu':
        thursdays(message)
    elif current_time == 'Fri':
        fridays(message)
    elif current_time == 'Sat':
        saturdays(message)
    else:
        bot.send_animation(message.chat.id, err_gf)


@bot.message_handler(commands=['nextday'])
def nextday(message):

    err_gf = open('data/err.gif', 'rb')

    if day_plus_one == 'Mon':
        mondays(message)
    elif day_plus_one == 'Tue':
        tuesdays(message)
    elif day_plus_one == 'Wed':
        wednesdays(message)
    elif day_plus_one == 'Thu':
        thursdays(message)
    elif day_plus_one == 'Fri':
        fridays(message)
    elif day_plus_one == 'Sat':
        saturdays(message)
    else:
        bot.send_animation(message.chat.id, err_gf)


@bot.message_handler(commands=['week'])
def send_week(message):
    weekn = open('data/weekn.png', 'rb')
    bot.send_document(message.chat.id, weekn)


@bot.message_handler(commands=['Расписание_звонков'])
def zvonok(message):
    zvonokp = open('data/raspisanie.jpg', 'rb')
    bot.send_photo(message.chat.id, zvonokp)


# Рассылка

@bot.message_handler(commands=['secret'])
def mess(message):
    for user in joinedUsers:
        bot.send_message(user, message.text[message.text.find(' '):])


@bot.message_handler(func=lambda message: True)
def echo_all(message):
    bot.reply_to(message, message.text)



def send_msg(text, id):
    try:
        bot.send_message(id, text)
    except Exception as e:
        logging.info(e)
        send_msg(text, id)


bot.polling(non_stop=True, interval=0)
