import telebot
import openpyxl
import datetime
import sqlite3
import os

TOKEN = os.getenv("TOKEN")
conn = sqlite3.connect('data.db')
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS users
             (chat_id INTEGER PRIMARY KEY, group_number INTEGER, department TEXT)''')

conn.commit()
conn.close()

# Порядковый номер предмета к расписанию звонков
# На Понедельник. Вторник
number_to_time_dict_mon = {'1': '8.00 - 9.20', '2': '9.35 - 10.55', '3': '11.10 - 12-30',
                           '4': '13.40 - 15.00', '5': '15.15 - 16.35', '6': '16.50 - 18.10'}
number_to_time_dict_wed = {'1': '8.00 - 9.30', '2': '9.45 - 11.15', '3': '11.30 - 13.00',
                           '4': '13.15 - 14.45', '5': '15.00 - 16.30', '6': '16.40 - 18.10'}
number_to_time_dict_sat = {'1': '8.00 - 9.20', '2': '9.30 - 10.50', '3': '11.00 - 12.20',
                           '4': '12.40 - 14.00', '5': '14.10 - 15.30', '6': '15.40 - 17.00'}

# Create the Telegram bot object
bot = telebot.TeleBot(TOKEN)


# Define the start command


@bot.message_handler(commands=['chgroup'])
def change_group(message):
    try:
        conn = sqlite3.connect('data.db')
        c = conn.cursor()
        c.execute('SELECT group_number FROM users WHERE chat_id = ?', (message.chat.id,))
        result = c.fetchone()
        print('Группа из БД', result)
        # The user has not entered their group number yet
        group_number = message.text.split(" ", 1)[1]  # Extract the group number from the message
        if not group_number.isdigit():
            bot.send_message(message.chat.id, "Неправильный номер группы, например 'Группа 2230'")
        else:
            print('Введенный номер группы ', group_number)
            c.execute('INSERT INTO users (chat_id, group_number) VALUES (?, ?) ON CONFLICT (chat_id) DO UPDATE SET '
                      'group_number = ?', (message.chat.id, group_number, group_number))
            conn.commit()
            conn.close()
            bot.send_message(message.chat.id, f"Группа успешно изменена на {group_number}")
    except IndexError:
        bot.send_message(message.chat.id, "Введите /chgroup №группы, например '/chgroup 2230'")


@bot.message_handler(commands=['about'])
def send_welcome(message):
    bot.reply_to(message,
                 'Привет! Я помогу тебе узнать расписание занятий для твоей группы. Для смены группы введи /chgroup '
                 '№Группы, например /chgroup 1234. Если выбрано не то отделение, то введи /start и выбери нужное')


@bot.message_handler(commands=['get_user_id'])
def send_user_id(message):
    user_id = message.from_user.id
    bot.send_message(user_id, f"Пользовательский номер: {user_id}")


@bot.message_handler(commands=['start'])
def send_welcome(message):
    bot.reply_to(message,
                 'Привет! Я помогу тебе узнать расписание занятий для твоей группы. Для смены группы введи /chgroup '
                 '№Группы, например /chgroup 1234. Больше информации в /about. Выбери отделение:',
                 reply_markup=get_department_keyboard(message))



def get_department_keyboard(message):
    markup = telebot.types.ReplyKeyboardMarkup(row_width=1)
    itembtn1 = telebot.types.KeyboardButton('Педагогическое')
    itembtn2 = telebot.types.KeyboardButton('Технологическое')
    itembtn3 = telebot.types.KeyboardButton('Строительное')
    markup.add(itembtn1, itembtn2, itembtn3)
    return markup


# Define the message handler
@bot.message_handler(func=lambda message: True)
def send_group_request_message(message):
    if message.text.startswith("Группа"):
        group_handler(message)
        print('Группа ', message.text)
    elif message.text in ['Сегодня', 'Завтра', 'Неделя']:
        group_handler(message)
        print('День ', message.text)
    elif message.text in ['Педагогическое', 'Технологическое', 'Строительное']:
        department_listener(message)
        print('Отделение ', message.text)
    else:
        bot.send_message(message.chat.id, 'Неизвестное сообщение')


def department_listener(message):
    # Save the selected department to the database
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute(
        'INSERT INTO users (chat_id, department) VALUES (?, ?) ON CONFLICT (chat_id) DO UPDATE SET department = ?',
        (message.chat.id, message.text, message.text))
    conn.commit()
    conn.close()
    bot.reply_to(message,
                 'Введите номер группы с помощью сообщения Группа №группы, например "Группа 1234".')


def group_handler(message):
    # Check if the user has already entered their group number
    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    c.execute('SELECT group_number FROM users WHERE chat_id = ?', (message.chat.id,))
    result = c.fetchone()
    print('Группа из БД', result)
    #print("Бот получает сообщение", message.text.split(" ", 1)[1])

    if result is None or result[0] is None:
        # The user has not entered their group number yet
        group_number = message.text.split(" ", 1)[1]  # Extract the group number from the message
        print('Введеный номер группы ', group_number)
        c.execute('INSERT INTO users (chat_id, group_number) VALUES (?, ?) ON CONFLICT (chat_id) DO UPDATE SET group_number = ?', (message.chat.id, group_number, group_number))
        conn.commit()
        conn.close()
    else:
        group_number = result[0]
        print('если номер имеется то, он = ', group_number)

    process_day_choice(message, group_number)
    try:
        conn = sqlite3.connect('data.db')
        c = conn.cursor()
        c.execute('SELECT group_number FROM users WHERE chat_id = ?', (message.chat.id,))
        result = c.fetchone()
        print('Группа из БД', result)
        # print("Бот получает сообщение", message.text.split(" ", 1)[1])
        if result is None or result[0] is None:
            # The user has not entered their group number yet
            group_number = message.text.split(" ", 1)[1]  # Extract the group number from the message
            if not group_number.isdigit():
                bot.send_message(message.chat.id, "Неправильный номер группы, например 'Группа 2230'")
            else:
                print('Введенный номер группы ', group_number)
                c.execute('INSERT INTO users (chat_id, group_number) VALUES (?, ?) ON CONFLICT (chat_id) DO UPDATE SET '
                          'group_number = ?', (message.chat.id, group_number, group_number))
                conn.commit()
                conn.close()
                process_day_choice(message, group_number)
        else:
            group_number = result[0]
            print('если номер имеется то, он = ', group_number)
            #process_day_choice(message, group_number)
        # process_day_choice(message, group_number)
    except IndexError:
        bot.send_message(message.chat.id, 'Введите номер группы с помощью слова Группа №группы, например Группа 2230')


# Define the day choice handler
def process_day_choice(message, group_number):
    now = None
    # Get the current day of the week
    now = datetime.datetime.now()

    # Get the day of the week for today and tomorrow
    day_names_dict = {'Monday': 'Понедельник', 'Tuesday': 'Вторник', 'Wednesday': 'Среда', 'Thursday': 'Четверг',
                      'Friday': 'Пятница', 'Saturday': 'Суббота', 'Sunday': 'Воскресенье'}

    today_day_name = now.strftime('%A')
    today_day_name_ru = day_names_dict.get(today_day_name, today_day_name)

    tomorrow = now + datetime.timedelta(days=1)
    tomorrow_day_name = tomorrow.strftime('%A')
    tomorrow_day_name_ru = day_names_dict.get(tomorrow_day_name, tomorrow_day_name)

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1)
    itembtn1 = telebot.types.KeyboardButton('Сегодня')
    itembtn2 = telebot.types.KeyboardButton('Завтра')
    itembtn3 = telebot.types.KeyboardButton('Неделя')
    markup.add(itembtn1, itembtn2, itembtn3)
    msg = bot.send_message(message.chat.id,
                           f'Хотите посмотреть расписание на сегодня - {today_day_name_ru}, на завтра - {tomorrow_day_name_ru} или на всю неделю?',
                           reply_markup=markup)

    day_choice = message.text.lower()

    if day_choice == 'сегодня':
        day_name_ru = today_day_name_ru
        days = [day_name_ru]
        process_schedule(message, group_number, days)
    elif day_choice == 'завтра':
        day_name_ru = tomorrow_day_name_ru
        days = [day_name_ru]
        process_schedule(message, group_number, days)
    elif day_choice == 'неделя':
        day_name_ru = list(day_names_dict.values())
        days = day_name_ru
        process_schedule(message, group_number, days)


def process_schedule(message, group_number, days):
    # Find the row with the group number
    def get_user_department(chat_id):
        conn = sqlite3.connect('data.db')
        c = conn.cursor()
        c.execute('SELECT department FROM users WHERE chat_id = ?', (chat_id,))
        result = c.fetchone()
        conn.close()
        if result is not None:
            return result[0]
        return None

    department = get_user_department(message.chat.id)
    try:
        if department == 'Педагогическое':
            workbook = openpyxl.load_workbook('ped.xlsx')
            sheet = workbook['1']
        elif department == 'Технологическое':
            workbook = openpyxl.load_workbook('tex.xlsx')
            sheet = workbook['1']
        elif department == 'Строительное':
            workbook = openpyxl.load_workbook('str.xlsx')
            sheet = workbook['1']
        else:
            bot.send_message(message.chat.id, 'Отделение не определено.')
            return
    except FileNotFoundError:
        bot.send_message(message.chat.id, "Не удалось найти файл с расписанием на сервере")
        print('Не удалось найти файл с расписанием на сервере')
        send_welcome(message)
        return

    group_row = None
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == f'Группа - {group_number}':
            group_row = row[0].row
            break
        elif row[0].value == f'Группа -{group_number}':
            group_row = row[0].row
            break

    # Send the class schedule for the chosen day(s) of the week for the given group
    if group_row is not None:
        schedule = f'Расписание занятий для группы - {group_number}:\n'
        for day_name_ru in days:
            # Find the column with the chosen day of the week
            day_col = None
            for col in sheet.iter_cols(min_row=7, max_row=7, min_col=2, max_col=sheet.max_column):
                if col[0].value == day_name_ru:
                    day_col = col[0].column
                    break
            if day_col is not None:
                schedule += f'\n{day_name_ru}:\n'
                item_number = 0
                for row in sheet.iter_rows(min_row=group_row + 3, max_row=group_row + 8, min_col=day_col,
                                           max_col=day_col + 2):
                    lesson_number = row[0].value
                    subject_name = row[1].value
                    item_number += 1
                    if day_name_ru in ['Понедельник', 'Вторник']:
                        number_to_time_dict = number_to_time_dict_mon
                    elif day_name_ru in ['Среда', 'Четверг', 'Пятница']:
                        number_to_time_dict = number_to_time_dict_wed
                    else:
                        number_to_time_dict = number_to_time_dict_sat
                    number_to_time = number_to_time_dict.get(str(item_number), str(item_number))
                    if lesson_number is not None:
                        schedule += f'{item_number}. {number_to_time}. {lesson_number}. {subject_name}\n'
            else:
                schedule += f'Расписание занятий на {day_name_ru} не найдено.\n'
        bot.send_message(message.chat.id, schedule)
    else:
        bot.send_message(message.chat.id, f'Группа - {group_number} не найдена в расписании.')


# Start the bot
bot.polling()
